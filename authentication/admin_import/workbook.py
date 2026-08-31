"""Lectura/escritura de workbooks openpyxl."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook

from .headers import TEMPLATE_HEADERS


def build_template_workbook(module: str) -> BytesIO:
    headers = TEMPLATE_HEADERS[module]
    wb = Workbook()
    ws = wb.active
    ws.title = module
    ws.append(headers)
    # Fila de ejemplo (comentarios en celdas vía segunda fila vacía opcional)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def read_rows(file_obj) -> tuple[list[str], list[dict[str, Any]]]:
    """
    Devuelve (headers, rows) donde cada row es dict header→valor (str limpio).
    """
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        raw_headers = next(rows_iter)
    except StopIteration:
        return [], []

    headers = [str(h).strip() if h is not None else "" for h in raw_headers]
    data: list[dict[str, Any]] = []
    for idx, raw in enumerate(rows_iter, start=2):
        if raw is None or all(c is None or str(c).strip() == "" for c in raw):
            continue
        row: dict[str, Any] = {"_row": idx}
        for i, key in enumerate(headers):
            if not key:
                continue
            val = raw[i] if i < len(raw) else None
            if val is None:
                row[key] = ""
            else:
                row[key] = val
        data.append(row)
    return headers, data


def cell_str(row: dict[str, Any], key: str) -> str:
    val = row.get(key, "")
    if val is None:
        return ""
    return str(val).strip()


def cell_bool(row: dict[str, Any], key: str, default: bool = False) -> bool:
    raw = cell_str(row, key).lower()
    if not raw:
        return default
    return raw in {"1", "true", "yes", "si", "sí", "y", "x"}
