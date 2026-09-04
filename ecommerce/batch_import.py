"""
Carga masiva de productos (CSV / Excel) para inventario de tienda.

Límites Plan Free (0.1 CPU / 512 MB): archivo ≤ 2 MB, ≤ 200 filas.
Procesa con openpyxl read_only / csv.DictReader y bulk_create/update (batch_size=50).
"""

from __future__ import annotations

import csv
import io
import logging
import os
from decimal import Decimal
from typing import Any
from zipfile import BadZipFile

from django.db import transaction
from django.http import HttpResponse
from django.utils.text import slugify
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from authentication.admin_import.headers import TEMPLATE_HEADERS, headers_match_module
from authentication.admin_import.importers import ImportResult
from authentication.admin_import.parsing import (
    RowImportError,
    parse_bool,
    parse_optional_decimal,
    parse_optional_int,
)
from authentication.admin_import.workbook import cell_str

logger = logging.getLogger(__name__)

PRODUCT_HEADERS = list(TEMPLATE_HEADERS["products"])

# Plan Free (Render): límites estrictos para evitar OOM / timeouts.
MAX_UPLOAD_BYTES = int(os.getenv("SHOP_BATCH_MAX_BYTES", str(2 * 1024 * 1024)))
MAX_BATCH_ROWS = int(os.getenv("SHOP_BATCH_MAX_ROWS", "200"))
DB_BATCH_SIZE = int(os.getenv("SHOP_BATCH_DB_SIZE", "50"))

INVALID_FILE_MESSAGE = (
    "El archivo no es un CSV o Excel válido (.csv, .xlsx)."
)


def build_product_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "products"
    ws.append(PRODUCT_HEADERS)
    ws.append(
        [
            "Camiseta ejemplo",
            "SKU-001",
            "Descripción larga",
            "Resumen corto",
            "Ropa",
            "Camisetas",
            "45000",
            "55000",
            "10",
            "https://example.com/foto.jpg",
            "false",
            "true",
        ]
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_product_template_csv() -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(PRODUCT_HEADERS)
    writer.writerow(
        [
            "Camiseta ejemplo",
            "SKU-001",
            "Descripción larga",
            "Resumen corto",
            "Ropa",
            "Camisetas",
            "45000",
            "55000",
            "10",
            "https://example.com/foto.jpg",
            "false",
            "true",
        ]
    )
    return buf.getvalue().encode("utf-8-sig")


def template_http_response(*, fmt: str = "xlsx") -> HttpResponse:
    fmt = (fmt or "xlsx").lower().strip()
    if fmt == "csv":
        payload = build_product_template_csv()
        response = HttpResponse(payload, content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = (
            'attachment; filename="chever_plantilla_productos.csv"'
        )
        return response
    payload = build_product_template_xlsx()
    response = HttpResponse(
        payload,
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = (
        'attachment; filename="chever_plantilla_productos.xlsx"'
    )
    response["X-Template-Headers"] = ",".join(PRODUCT_HEADERS)
    return response


def _is_blank_cell(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def read_upload_rows(upload) -> tuple[list[str], list[dict[str, Any]]]:
    """
    Lee CSV o Excel en modo streaming / read_only.
    Lanza ValueError si el formato es inválido o se exceden filas.
    """
    name = str(getattr(upload, "name", "") or "").lower()
    size = int(getattr(upload, "size", 0) or 0)
    if size > MAX_UPLOAD_BYTES:
        raise ValueError(
            f"El archivo supera el máximo de {MAX_UPLOAD_BYTES // (1024 * 1024)} MB "
            f"permitido en el plan actual."
        )

    if name.endswith(".csv"):
        return _read_csv_rows(upload)
    if name.endswith((".xlsx", ".xlsm", ".xls")):
        return _read_xlsx_rows(upload)
    raise ValueError(INVALID_FILE_MESSAGE)


def _read_csv_rows(upload) -> tuple[list[str], list[dict[str, Any]]]:
    raw = upload.read()
    if isinstance(raw, bytes):
        text = raw.decode("utf-8-sig", errors="replace")
    else:
        text = str(raw)
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return [], []
    headers = [str(h).strip() for h in reader.fieldnames if h is not None]
    data: list[dict[str, Any]] = []
    for idx, row in enumerate(reader, start=2):
        if not row or all(_is_blank_cell(v) for v in row.values()):
            continue
        if len(data) >= MAX_BATCH_ROWS:
            raise ValueError(
                f"El archivo supera el máximo de {MAX_BATCH_ROWS} filas "
                "permitido en el plan Free."
            )
        item = {"_row": idx}
        for key in headers:
            item[key] = row.get(key, "")
        data.append(item)
    return headers, data


def _read_xlsx_rows(upload) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        wb = load_workbook(upload, read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, KeyError) as exc:
        raise ValueError(INVALID_FILE_MESSAGE) from exc
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        raw_headers = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in raw_headers]
    data: list[dict[str, Any]] = []
    for idx, raw in enumerate(rows_iter, start=2):
        if raw is None or all(_is_blank_cell(c) for c in raw):
            continue
        if len(data) >= MAX_BATCH_ROWS:
            wb.close()
            raise ValueError(
                f"El archivo supera el máximo de {MAX_BATCH_ROWS} filas "
                "permitido en el plan Free."
            )
        row: dict[str, Any] = {"_row": idx}
        for i, key in enumerate(headers):
            if not key:
                continue
            val = raw[i] if i < len(raw) else None
            row[key] = "" if val is None else val
        data.append(row)
    wb.close()
    return headers, data


def _allocate_slugs(org, names: list[str]) -> list[str]:
    """Asigna slugs únicos (una query de existentes + set en memoria)."""
    from ecommerce.models import Product

    used = set(
        Product.objects.filter(organization=org).values_list("slug", flat=True)
    )
    out: list[str] = []
    for name in names:
        base = slugify(name)[:200] or "producto"
        candidate = base
        n = 1
        while candidate in used:
            candidate = f"{base}-{n}"[:220]
            n += 1
        used.add(candidate)
        out.append(candidate)
    return out


def import_products_batch(*, rows: list[dict], organization, user) -> ImportResult:
    """
    Importa productos con created_by=user.
    Valida fila a fila; inserta/actualiza con bulk_* (batch_size=50).
    """
    from ecommerce.models import Category, Product, SubCategory

    result = ImportResult()
    to_create: list[Product] = []
    to_update: list[Product] = []
    update_fields = [
        "name",
        "description",
        "short_description",
        "sku",
        "price_cop",
        "compare_at_price_cop",
        "stock",
        "image_url",
        "is_featured",
        "is_published",
        "category",
        "subcategory",
        "updated_at",
    ]

    # Precarga SKUs existentes de la org (una query).
    skus = [
        cell_str(row, "sku")
        for row in rows
        if cell_str(row, "sku")
    ]
    existing_by_sku = {
        p.sku: p
        for p in Product.objects.filter(organization=organization, sku__in=skus)
    }

    category_cache: dict[str, Category] = {}
    subcategory_cache: dict[tuple[str, str], SubCategory] = {}
    pending_creates: list[dict[str, Any]] = []
    create_names: list[str] = []

    for row in rows:
        r = int(row.get("_row") or 0)
        try:
            row_errors: list[tuple[str | None, str]] = []
            name = cell_str(row, "name")
            if not name:
                row_errors.append(("name", "El campo 'name' es obligatorio y está vacío."))

            price = None
            try:
                price = parse_optional_decimal(row.get("price_cop"), "price_cop")
            except RowImportError as exc:
                row_errors.append((exc.field, str(exc)))
            if price is None and not any(f == "price_cop" for f, _ in row_errors):
                row_errors.append(
                    ("price_cop", "El campo 'price_cop' es obligatorio y está vacío.")
                )
            if price is not None and price < 0:
                row_errors.append(("price_cop", "El precio no puede ser negativo."))

            compare_at = None
            try:
                compare_at = parse_optional_decimal(
                    row.get("compare_at_price_cop"), "compare_at_price_cop"
                )
            except RowImportError as exc:
                row_errors.append((exc.field, str(exc)))

            stock = 0
            try:
                stock = parse_optional_int(row.get("stock"), "stock", default=0) or 0
                if stock < 0:
                    row_errors.append(("stock", "El campo 'stock' no puede ser negativo."))
            except RowImportError as exc:
                row_errors.append((exc.field, str(exc)))

            is_featured = False
            is_published = True
            try:
                is_featured = parse_bool(row.get("is_featured"), "is_featured")
            except RowImportError as exc:
                row_errors.append((exc.field, str(exc)))
            try:
                is_published = parse_bool(
                    row.get("is_published"), "is_published", default=True
                )
            except RowImportError as exc:
                row_errors.append((exc.field, str(exc)))

            if row_errors:
                for field, msg in row_errors:
                    result.add_error(r, msg, field)
                continue

            cat_name = cell_str(row, "category")
            sub_name = cell_str(row, "subcategory")
            category = None
            subcategory = None
            if cat_name:
                cache_key = slugify(cat_name)[:140] or "general"
                category = category_cache.get(cache_key)
                if category is None:
                    category, _ = Category.objects.get_or_create(
                        organization=organization,
                        slug=cache_key,
                        defaults={"name": cat_name},
                    )
                    if category.name != cat_name:
                        category.name = cat_name
                        category.save(update_fields=["name", "updated_at"])
                    category_cache[cache_key] = category
            if category and sub_name:
                sub_key = (category.slug, slugify(sub_name)[:140] or "sub")
                subcategory = subcategory_cache.get(sub_key)
                if subcategory is None:
                    subcategory, _ = SubCategory.objects.get_or_create(
                        organization=organization,
                        category=category,
                        slug=sub_key[1],
                        defaults={"name": sub_name},
                    )
                    subcategory_cache[sub_key] = subcategory

            sku = cell_str(row, "sku")
            payload = {
                "name": name,
                "description": cell_str(row, "description"),
                "short_description": cell_str(row, "short_description")[:300],
                "sku": sku,
                "price_cop": price,
                "compare_at_price_cop": compare_at,
                "stock": stock,
                "image_url": cell_str(row, "image_url"),
                "is_featured": is_featured,
                "is_published": is_published,
                "category": category,
                "subcategory": subcategory,
            }
            existing = existing_by_sku.get(sku) if sku else None
            if existing:
                for k, v in payload.items():
                    setattr(existing, k, v)
                to_update.append(existing)
            else:
                create_names.append(name)
                pending_creates.append(payload)
        except Exception as exc:  # noqa: BLE001
            result.add_error(r, f"Error en la fila {r}: {exc}", None)

    slugs = _allocate_slugs(organization, create_names)
    for payload, slug in zip(pending_creates, slugs):
        to_create.append(
            Product(
                organization=organization,
                created_by=user,
                slug=slug,
                **payload,
            )
        )

    with transaction.atomic():
        if to_create:
            Product.objects.bulk_create(to_create, batch_size=DB_BATCH_SIZE)
            result.created = len(to_create)
        if to_update:
            Product.objects.bulk_update(
                to_update, fields=update_fields, batch_size=DB_BATCH_SIZE
            )
            result.updated = len(to_update)

    return result


def assert_product_headers(headers: list[str]) -> None:
    if not headers_match_module(headers, "products"):
        raise ValueError(
            "La plantilla subida no corresponde a productos. "
            "Descarga la plantilla oficial e inténtalo de nuevo."
        )
